import os
import uuid
import base64
import json
import re

import fitz  # pymupdf
import requests
import pandas as pd
from flask import Flask, render_template, request, redirect, url_for, flash, send_from_directory
from werkzeug.utils import secure_filename
from dotenv import load_dotenv

load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "dev-secret-key")
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100 MB

GROQ_API_KEY  = os.getenv("GROQ_API_KEY", "")
GROQ_MODEL    = os.getenv("GROQ_MODEL", "meta-llama/llama-4-scout-17b-16e-instruct")
UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), "uploads")
OUTPUT_FOLDER = os.path.join(os.path.dirname(__file__), "outputs")
ALLOWED_EXTS  = {"pdf", "jpg", "jpeg", "png"}

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

PROMPT = """
You are a data extraction AI. Read the document and extract ALL data fields.

STRICT RULES:
1. Return ONLY a valid JSON array — no markdown, no backticks, no explanation.
2. Each item in the array must be a flat key-value object.
3. If the document has a table with multiple rows, each row = one object.
4. If it's a single form/card, return an array with one object.
5. Keys must be readable column names like "Name", "Date", "Amount".
6. No nested objects. Values must be strings or numbers only.

Example:
[
  {"Name": "Ravi", "Age": 22, "City": "Mumbai"},
  {"Name": "Priya", "Age": 25, "City": "Pune"}
]

Now extract all data from this document and return the JSON array only.
""".strip()


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTS


def pdf_to_base64_images(pdf_path, max_pages=3):
    """Convert PDF pages to compressed base64 JPEG images."""
    doc = fitz.open(pdf_path)
    images = []
    total = min(len(doc), max_pages)
    for i in range(total):
        page = doc[i]
        mat = fitz.Matrix(1.2, 1.2)   # ~86 DPI — readable but small
        pix = page.get_pixmap(matrix=mat)
        # Save as JPEG with compression (quality 70)
        img_bytes = pix.tobytes("jpeg", jpg_quality=70)
        images.append(base64.b64encode(img_bytes).decode("utf-8"))
    doc.close()
    return images


def _clean_json(text):
    """Fix common JSON issues produced by LLMs."""
    # Remove trailing commas before ] or }
    text = re.sub(r",\s*([\]\}])", r"\1", text)
    # Replace smart/curly quotes with straight quotes
    text = text.replace("\u201c", '"').replace("\u201d", '"')
    text = text.replace("\u2018", "'").replace("\u2019", "'")
    # Remove control characters except newline/tab
    text = re.sub(r"[\x00-\x08\x0b-\x0c\x0e-\x1f]", "", text)
    # Fix unescaped newlines inside string values
    text = re.sub(r'(?<="):([^"]*)\n([^"]*?)(?=")', r':\1 \2', text)
    return text


def _extract_objects(text):
    """Extract individual JSON objects when full array parse fails."""
    records = []
    for match in re.finditer(r"\{[^{}]+\}", text):
        try:
            obj = json.loads(_clean_json(match.group(0)))
            if isinstance(obj, dict):
                records.append(obj)
        except Exception:
            continue
    if not records:
        raise ValueError(
            "Could not parse structured JSON from the AI response. "
            "Try a clearer document or a different file."
        )
    return records


def call_groq(file_path, file_ext):
    """Send file to Groq vision model and return extracted JSON records."""
    if not GROQ_API_KEY:
        raise ValueError("GROQ_API_KEY is not set in .env file.")

    # Build image content blocks
    image_blocks = []

    if file_ext == "pdf":
        pages = pdf_to_base64_images(file_path)
        for b64 in pages:
            image_blocks.append({
                "type": "image_url",
                "image_url": {"url": f"data:image/jpeg;base64,{b64}"}
            })
    else:
        mime = "image/jpeg" if file_ext in ("jpg", "jpeg") else "image/png"
        with open(file_path, "rb") as f:
            b64 = base64.b64encode(f.read()).decode("utf-8")
        image_blocks.append({
            "type": "image_url",
            "image_url": {"url": f"data:{mime};base64,{b64}"}
        })

    # Add the prompt text block
    image_blocks.append({"type": "text", "text": PROMPT})

    payload = {
        "model": GROQ_MODEL,
        "messages": [{"role": "user", "content": image_blocks}],
        "temperature": 0.1,
        "max_tokens": 4096,
    }

    headers = {
        "Authorization": f"Bearer {GROQ_API_KEY}",
        "Content-Type": "application/json",
    }

    resp = requests.post(
        "https://api.groq.com/openai/v1/chat/completions",
        headers=headers,
        json=payload,
        timeout=60,
    )

    if resp.status_code == 429:
        raise ValueError("Groq rate limit hit. Wait a moment and try again.")
    if resp.status_code == 401:
        raise ValueError("Invalid GROQ_API_KEY. Check your .env file.")
    if not resp.ok:
        msg = resp.json().get("error", {}).get("message", resp.text[:200])
        raise ValueError(f"Groq API error {resp.status_code}: {msg}")

    raw = resp.json()["choices"][0]["message"]["content"].strip()

    # Strip markdown fences if present
    fence = re.search(r"```(?:json)?\s*([\s\S]+?)\s*```", raw)
    if fence:
        raw = fence.group(1).strip()

    # Find JSON array
    arr = re.search(r"\[[\s\S]*\]", raw)
    if arr:
        raw = arr.group(0)

    # Clean common JSON issues from LLMs
    raw = _clean_json(raw)

    try:
        records = json.loads(raw)
    except json.JSONDecodeError:
        # Last resort: extract individual objects and parse one by one
        records = _extract_objects(raw)

    if isinstance(records, dict):
        records = [records]
    return records


def make_excel(records, output_path):
    """Convert records to a styled Excel file."""
    df = pd.DataFrame(records)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Extracted Data")
        ws = writer.sheets["Extracted Data"]

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        header_fill = PatternFill("solid", fgColor="1B4F72")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin = Side(style="thin", color="BDC3C7")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        odd_fill  = PatternFill("solid", fgColor="EBF5FB")
        even_fill = PatternFill("solid", fgColor="FFFFFF")

        # Header row
        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        ws.row_dimensions[1].height = 28

        # Data rows
        for row in range(2, len(df) + 2):
            fill = odd_fill if row % 2 == 0 else even_fill
            ws.row_dimensions[row].height = 20
            for col in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = fill
                cell.alignment = Alignment(vertical="center")
                cell.border = border

        # Column widths
        for i, col_name in enumerate(df.columns, 1):
            max_len = max(len(str(col_name)), df[col_name].astype(str).str.len().max())
            ws.column_dimensions[get_column_letter(i)].width = min(max(max_len + 4, 12), 50)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions


# ── Routes ────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/upload", methods=["POST"])
def upload():
    if not GROQ_API_KEY:
        flash("GROQ_API_KEY not set in .env file.", "error")
        return redirect(url_for("index"))

    file = request.files.get("document")

    if not file or file.filename == "":
        flash("No file selected.", "error")
        return redirect(url_for("index"))

    if not allowed_file(file.filename):
        flash("Only PDF, JPG, JPEG, PNG files are allowed.", "error")
        return redirect(url_for("index"))

    # Save file
    ext = file.filename.rsplit(".", 1)[1].lower()
    filename = f"{uuid.uuid4().hex}.{ext}"
    upload_path = os.path.join(UPLOAD_FOLDER, filename)
    file.save(upload_path)

    # Call Groq
    try:
        records = call_groq(upload_path, ext)
    except ValueError as e:
        flash(str(e), "error")
        return redirect(url_for("index"))
    except Exception as e:
        flash(f"Extraction failed: {e}", "error")
        return redirect(url_for("index"))
    finally:
        if os.path.exists(upload_path):
            os.remove(upload_path)

    if not records:
        flash("No data could be extracted from this document.", "warning")
        return redirect(url_for("index"))

    # Make Excel
    excel_name = f"data_{uuid.uuid4().hex[:8]}.xlsx"
    excel_path = os.path.join(OUTPUT_FOLDER, excel_name)
    try:
        make_excel(records, excel_path)
    except Exception as e:
        flash(f"Excel creation failed: {e}", "error")
        return redirect(url_for("index"))

    return render_template(
        "result.html",
        records=records,
        columns=list(records[0].keys()),
        excel_name=excel_name,
        original_name=secure_filename(file.filename),
        total=len(records),
    )


@app.route("/download/<filename>")
def download(filename):
    safe = secure_filename(filename)
    path = os.path.join(OUTPUT_FOLDER, safe)
    if not os.path.isfile(path):
        flash("File not found.", "error")
        return redirect(url_for("index"))
    return send_from_directory(OUTPUT_FOLDER, safe, as_attachment=True)


if __name__ == "__main__":
    app.run(debug=True)